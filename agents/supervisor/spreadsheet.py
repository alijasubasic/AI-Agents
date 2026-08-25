"""Writing the morning brief out as a spreadsheet.

Same pattern as every other external dependency here: one `Protocol`, and
implementations behind it. The difference is that both implementations are
real — CSV is not a mock of XLSX, it is the zero-dependency option.

CSV is the default so `make demo` works on a clone with nothing extra
installed. XLSX produces one workbook with a tab per sheet and needs the
optional `xlsx` extra.
"""

from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Protocol

from pydantic import BaseModel, Field


class Sheet(BaseModel):
    """One tab: a header row and the rows under it."""

    name: str
    columns: list[str]
    rows: list[list[str]] = Field(default_factory=list)

    @property
    def safe_name(self) -> str:
        """A filename-safe version of the sheet name."""
        return re.sub(r"[^a-z0-9]+", "-", self.name.lower()).strip("-")


class SpreadsheetWriter(Protocol):
    """Writes a set of sheets somewhere."""

    suffix: str

    def write(self, sheets: list[Sheet], destination: Path) -> list[Path]: ...


class CsvWorkbook:
    """One CSV file per sheet, in a directory. No dependencies.

    Excel on a German locale opens semicolon-separated files more happily than
    comma-separated ones, so the delimiter is configurable and the default
    matches the audience this brief is written for.
    """

    suffix = ".csv"

    def __init__(self, delimiter: str = ";") -> None:
        self.delimiter = delimiter

    def write(self, sheets: list[Sheet], destination: Path) -> list[Path]:
        destination.mkdir(parents=True, exist_ok=True)
        written: list[Path] = []

        for index, sheet in enumerate(sheets, start=1):
            path = destination / f"{index:02d}-{sheet.safe_name}.csv"
            # utf-8-sig: without the BOM, Excel reads UTF-8 CSV as the local
            # code page and mangles every accented character.
            with path.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.writer(handle, delimiter=self.delimiter)
                writer.writerow(sheet.columns)
                writer.writerows(sheet.rows)
            written.append(path)

        return written


class XlsxWorkbook:
    """A single .xlsx with one tab per sheet. Needs the `xlsx` extra."""

    suffix = ".xlsx"

    def __init__(self, freeze_header: bool = True) -> None:
        self.freeze_header = freeze_header

    def write(self, sheets: list[Sheet], destination: Path) -> list[Path]:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter
        except ImportError as exc:  # pragma: no cover - optional dependency
            raise ImportError("XlsxWorkbook needs openpyxl: uv sync --extra xlsx") from exc

        path = destination if destination.suffix == ".xlsx" else destination / "brief.xlsx"
        path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        workbook.remove(workbook.active)

        for sheet in sheets:
            # Excel refuses sheet names over 31 characters.
            worksheet = workbook.create_sheet(sheet.name[:31])
            worksheet.append(sheet.columns)
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
            for row in sheet.rows:
                worksheet.append(row)

            if self.freeze_header:
                worksheet.freeze_panes = "A2"
            _autosize(worksheet, sheet, get_column_letter)

        workbook.save(path)
        return [path]


def _autosize(worksheet, sheet: Sheet, get_column_letter) -> None:
    """Widen columns to fit their content, within reason.

    Unwidened columns show as ### or truncate, which makes the brief look
    broken before anyone has read a word of it.
    """
    for index, column in enumerate(sheet.columns, start=1):
        longest = max(
            [len(str(column))]
            + [len(str(row[index - 1])) for row in sheet.rows if len(row) >= index]
        )
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(longest + 2, 10), 70)


def build_writer(fmt: str = "csv") -> SpreadsheetWriter:
    """Pick a writer by name. Unknown names fail loudly rather than silently."""
    writers = {"csv": CsvWorkbook, "xlsx": XlsxWorkbook}
    if fmt not in writers:
        raise ValueError(f"Unknown spreadsheet format {fmt!r}; choose one of {sorted(writers)}")
    return writers[fmt]()
